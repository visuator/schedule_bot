import openpyxl
import argparse
import faker
import datetime
from openpyxl import Workbook

IMPORT_PATTERN_PATH = "import_pattern.xlsx"
TIMES = [
    "8:00",
    "9:40",
    "13:10",
    "14:00",
    "16:10"
]

def copy_cell(source, target, row, column):
    source_cell = source.cell(row=row, column=column)
    target_cell = target.cell(row=row, column=column)
    target_cell.value = source_cell.value

def write_cell(target, row, column, value):
    cell = target.cell(row=row, column=column)
    cell.value = value

parser = argparse.ArgumentParser(
    prog="utility for the schedule_bot which generates fake schedule"
)
parser.add_argument("rows_count", help="amount of generated rows")
parser.add_argument("--output", "-o", default=datetime.datetime.now().strftime("%d-%m-%y"), help="output file name")

args = parser.parse_args()

fake = faker.Faker('ru_RU')

source = openpyxl.load_workbook(IMPORT_PATTERN_PATH, read_only=True).active

workbook = Workbook()
target = workbook.active

# copying of header row
for i in range(1, 8):
    copy_cell(source, target, 1, i)

for i in range(2, int(args.rows_count) + 1):
    write_cell(target, i, 1, i)
    write_cell(target, i, 2, i % 2 + 1)
    write_cell(target, i, 3, fake.day_of_week())
    write_cell(target, i, 4, fake.bs())
    write_cell(target, i, 5, fake.name())
    write_cell(target, i, 6, fake.random_number(digits=3))
    write_cell(target, i, 7, fake.random_element(TIMES))

workbook.save(filename=f"gen/{args.output}.xlsx")